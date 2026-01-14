"""
API Serverless pour extraire les données ZEFIX
Compatible avec Vercel Serverless Functions
Version corrigée - Extraction réelle des données
"""

from http.server import BaseHTTPRequestHandler
import json
import requests
from datetime import datetime, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import base64


class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            # Lire le body
            content_length = int(self.headers.get('Content-Length', 0))
            post_data = self.rfile.read(content_length) if content_length > 0 else b'{}'
            data = json.loads(post_data.decode('utf-8'))
            
            cantons = data.get('cantons', ['GE', 'VD'])
            days = data.get('days', 7)
            
            # Extraire les données
            entreprises = self.extract_zefix(cantons, days)
            
            # Créer le fichier Excel
            excel_data = self.create_excel(entreprises)
            
            # Encoder en base64 pour le retour
            excel_b64 = base64.b64encode(excel_data).decode('utf-8')
            
            # Préparer la réponse
            response = {
                'success': True,
                'count': len(entreprises),
                'filename': f'Nouvelles_Entreprises_{datetime.now().strftime("%Y-%m-%d")}.xlsx',
                'data': excel_b64
            }
            
            # Envoyer la réponse
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            
            self.wfile.write(json.dumps(response).encode('utf-8'))
            
        except Exception as e:
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            
            error_response = {
                'success': False,
                'error': str(e),
                'type': type(e).__name__
            }
            
            self.wfile.write(json.dumps(error_response).encode('utf-8'))
    
    def do_OPTIONS(self):
        """Handle CORS preflight"""
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
    
    def extract_zefix(self, cantons, days):
        """Extraire les données depuis l'API ZEFIX"""
        entreprises = []
        
        # Date de début (format YYYY-MM-DD)
        date_from = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
        
        # API ZEFIX publique - endpoint correct
        base_url = "https://www.zefix.admin.ch/ZefixPublicREST/api/v1/firm/search.json"
        
        for canton in cantons:
            print(f"Extraction canton {canton}...")
            
            # Paramètres de recherche
            params = {
                'name': '',  # Vide = toutes les entreprises
                'canton': canton,
                'activeOnly': 'false',
                'deleteDate': '',
                'registryOfCommerceId': '',
                'legalSeat': ''
            }
            
            try:
                # Requête à l'API
                response = requests.get(base_url, params=params, timeout=30)
                
                if response.status_code == 200:
                    data = response.json()
                    
                    # Parser les résultats
                    if 'list' in data:
                        for item in data['list']:
                            # Filtrer par date d'inscription
                            date_str = item.get('inscription', {}).get('date', '')
                            if date_str and date_str >= date_from:
                                # Extraire les informations
                                forme = self.get_forme_juridique(item.get('legalForm', ''))
                                
                                # Filtrer uniquement SA, Sàrl, EI
                                if forme in ['SA', 'Sàrl', 'EI']:
                                    entreprise = {
                                        'nom': item.get('name', ''),
                                        'forme_juridique': forme,
                                        'canton': canton,
                                        'ville': item.get('address', {}).get('city', ''),
                                        'npa': item.get('address', {}).get('swissZipCode', ''),
                                        'adresse': self.format_adresse(item.get('address', {})),
                                        'date_inscription': date_str,
                                        'uid': item.get('uid', ''),
                                        'numero_rc': item.get('chId', '')
                                    }
                                    
                                    entreprises.append(entreprise)
                    
                    print(f"Canton {canton}: {len([e for e in entreprises if e['canton'] == canton])} entreprises trouvées")
                                    
            except Exception as e:
                print(f"Erreur canton {canton}: {e}")
                continue
        
        # Trier par priorité
        entreprises = self.prioritize(entreprises)
        
        print(f"Total: {len(entreprises)} entreprises")
        
        return entreprises
    
    def format_adresse(self, address):
        """Formater l'adresse complète"""
        parts = []
        
        if address.get('street'):
            parts.append(address['street'])
        if address.get('houseNumber'):
            if parts:
                parts[-1] += f" {address['houseNumber']}"
            else:
                parts.append(address['houseNumber'])
        
        return ' '.join(parts) if parts else ''
    
    def get_forme_juridique(self, legal_form):
        """Convertir le code en forme juridique"""
        legal_form_str = str(legal_form).lower()
        
        if 'sa' in legal_form_str or 'aktiengesellschaft' in legal_form_str or '0106' in legal_form_str:
            return 'SA'
        elif 'sàrl' in legal_form_str or 'sarl' in legal_form_str or 'gmbh' in legal_form_str or '0107' in legal_form_str:
            return 'Sàrl'
        elif 'individuel' in legal_form_str or 'einzelfirma' in legal_form_str or '0108' in legal_form_str:
            return 'EI'
        else:
            return 'Autre'
    
    def prioritize(self, entreprises):
        """Prioriser par forme juridique"""
        priority_map = {'SA': 1, 'Sàrl': 2, 'EI': 3}
        
        return sorted(entreprises, key=lambda x: priority_map.get(x['forme_juridique'], 99))
    
    def create_excel(self, entreprises):
        """Créer le fichier Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Nouvelles Entreprises"
        
        # En-têtes
        headers = [
            'Priorité', 'Nom entreprise', 'Forme juridique', 'Canton', 
            'Ville', 'NPA', 'Adresse', 'Date publication', 
            'Téléphone', 'Email', 'Site web', 'LinkedIn', 
            'Statut', 'Notes', 'Date dernier contact', 'Numéro RC', 'UID'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Données
        for row_idx, ent in enumerate(entreprises, 2):
            priorite = 'Haute' if ent['forme_juridique'] == 'SA' else 'Moyenne' if ent['forme_juridique'] == 'Sàrl' else 'Basse'
            
            row_data = [
                priorite,
                ent['nom'],
                ent['forme_juridique'],
                ent['canton'],
                ent['ville'],
                ent['npa'],
                ent['adresse'],
                ent['date_inscription'],
                '',
                '',
                '',
                '',
                'Nouveau',
                '',
                '',
                ent['numero_rc'],
                ent['uid']
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                
                if col_idx == 1 and priorite in ['Haute', 'Moyenne', 'Basse']:
                    color_map = {'Haute': 'FFCCCC', 'Moyenne': 'FFE5CC', 'Basse': 'FFFFCC'}
                    cell.fill = PatternFill(start_color=color_map[priorite], 
                                           end_color=color_map[priorite], 
                                           fill_type='solid')
                    cell.font = Font(bold=True)
        
        # Largeurs colonnes
        column_widths = {
            'A': 12, 'B': 40, 'C': 15, 'D': 10, 'E': 20, 
            'F': 8, 'G': 35, 'H': 15, 'I': 18, 'J': 30, 
            'K': 35, 'L': 35, 'M': 12, 'N': 40, 'O': 18, 
            'P': 20, 'Q': 20
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Figer + filtres
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        
        # Sauvegarder en mémoire
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.read()
